# modules/products.py
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from modules.database import DatabaseManager
import logging

logger = logging.getLogger(__name__)


class ProductManager:
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager

    def create_product(self, product_data):
        """Создание нового изделия в базе данных"""
        logger.debug(f"[ИЗДЕЛИЯ] Создание изделия с данными: {product_data}")
        query = """
            INSERT INTO products (product_id, article, name)
            VALUES (?, ?, ?)
        """
        params = (
            product_data.get('product_id', ''),
            product_data.get('article', ''),
            product_data.get('name', '')
        )

        # Используем контекстный менеджер для корректного управления соединением
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, params)
            # Получаем ID последней вставленной записи через lastrowid
            # Это надежнее, чем отдельный запрос last_insert_rowid()
            # в другом соединении или после закрытия соединения.
            product_id = cursor.lastrowid
            conn.commit()  # Обязательно делаем commit, если используем контекстный менеджер напрямую

        logger.debug(f"[ИЗДЕЛИЯ] Изделие создано в БД с ID: {product_id}")
        # Убедимся, что мы возвращаем целое число, а не кортеж или None
        if product_id is None:
            logger.error("[ИЗДЕЛИЯ] Не удалось получить ID созданного изделия!")
            # Попробуем альтернативный способ - получить максимальный ID
            try:
                alt_id_result = self.db_manager.fetch_one("SELECT MAX(id) FROM products")
                product_id = alt_id_result[0] if alt_id_result and alt_id_result[0] is not None else None
                if product_id:
                    logger.warning(f"[ИЗДЕЛИЯ] Альтернативный способ дал ID: {product_id}")
                else:
                    logger.error("[ИЗДЕЛИЯ] Альтернативный способ тоже не дал результата.")
            except Exception as e:
                logger.error(f"[ИЗДЕЛИЯ] Ошибка альтернативного способа: {e}")

        return product_id

    def save_product_to_excel(self, product_id, file_path):
        """
        Сохранение изделия в Excel файл с форматированием.
        Использует openpyxl напрямую для обеспечения целостности файла.
        Также сохраняет операции и материалы в БД.
        """
        logger.info(f"[ИЗДЕЛИЯ_EXCEL] Начало сохранения изделия ID {product_id} в '{file_path}'")
        try:
            # --- 1. Получение данных из БД ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 1. Получение данных из БД")

            product_info = self.db_manager.fetch_one(
                "SELECT * FROM products WHERE id = ?", (product_id,)
            )

            if not product_info:
                logger.error(f"[ИЗДЕЛИЯ_EXCEL] Ошибка: Изделие с ID {product_id} не найдено в БД")
                return False

            logger.debug(f"[ИЗДЕЛИЯ_EXCEL] Информация об изделии получена: {product_info}")

            # --- 2. Получение операций из БД ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 2. Получение операций из БД")
            operations_raw = self.db_manager.fetch_all("""
                SELECT 
                    COALESCE(o.operation_name, ''),
                    COALESCE(o.quantity_measured, 0),
                    COALESCE(o.time_measured, 0.0),
                    COALESCE(o.time_per_unit, 0.0),
                    COALESCE(o.rate_per_minute, 0.0),
                    COALESCE(o.cost, 0.0),
                    COALESCE(e.name, ''),
                    COALESCE(o.approved_rate, '')
                FROM operations o
                LEFT JOIN employees e ON o.employee_id = e.id
                WHERE o.product_id = ?
                ORDER BY o.id
            """, (product_id,))

            # Преобразование кортежей в списки для возможности изменения (если нужно)
            operations = [list(op) for op in operations_raw]
            logger.debug(f"[ИЗДЕЛИЯ_EXCEL] Получено {len(operations) if operations else 0} операций")

            # --- 3. Получение материалов из БД ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 3. Получение материалов из БД")
            materials_raw = self.db_manager.fetch_all("""
                SELECT 
                    COALESCE(m.name, ''),
                    COALESCE(pm.length, 0.0),
                    COALESCE(pm.width, 0.0),
                    COALESCE(pm.quantity, 0),
                    COALESCE(pm.cost, 0.0)
                FROM product_materials pm
                JOIN materials m ON pm.material_id = m.id
                WHERE pm.product_id = ?
                ORDER BY pm.id
            """, (product_id,))

            # Преобразование кортежей в списки
            materials = [list(mat) for mat in materials_raw]
            logger.debug(f"[ИЗДЕЛИЯ_EXCEL] Получено {len(materials) if materials else 0} материалов")

            # --- 4. Создание Excel файла с помощью openpyxl ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 4. Создание Excel файла с помощью openpyxl")

            # Убедимся, что директория существует
            output_dir = os.path.dirname(file_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            logger.debug(f"[ИЗДЕЛИЯ_EXCEL] Директория для файла создана/проверена: {output_dir}")

            # Создаем новую книгу
            wb = Workbook()
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Создана новая книга Workbook()")

            # Удаляем дефолтный лист, если он есть
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Удалён дефолтный лист 'Sheet'")

            # --- 5. Создание и заполнение листа "Информация" ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 5. Создание и заполнение листа 'Информация'")

            ws_info = wb.create_sheet(title="Информация")
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Информация' создан")

            # Заголовки
            ws_info.append(['Поле', 'Значение'])
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Заголовки добавлены на лист 'Информация'")

            # Данные изделия
            info_data = [
                ["ID", str(product_info[1]) if product_info[1] is not None else ""],
                ["Артикул", str(product_info[2]) if product_info[2] is not None else ""],
                ["Название", str(product_info[3]) if product_info[3] is not None else ""]
            ]

            for row_data in info_data:
                ws_info.append(row_data)
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Данные изделия добавлены на лист 'Информация'")

            # Форматирование листа "Информация"
            self._format_info_sheet(ws_info)
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Информация' отформатирован")

            # --- 6. Создание и заполнение листа "Операции" ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 6. Создание и заполнение листа 'Операции'")

            if operations:
                ws_ops = wb.create_sheet(title="Операции")
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Операции' создан")

                # Заголовки для операций
                ops_headers = [
                    "Операция", "Кол-во по замерам", "Время замера (мин)",
                    "Время на 1 деталь (мин)", "Ставка (грн/мин)", "Стоимость (грн)",
                    "Сотрудник", "Утверждённая расценка"
                ]
                ws_ops.append(ops_headers)
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Заголовки добавлены на лист 'Операции'")

                # Данные операций
                for op_data in operations:
                    # Убедимся, что все данные преобразованы в строки или числа
                    row_to_append = [
                        str(op_data[0]),  # Операция
                        str(op_data[1]),  # Кол-во по замерам
                        f"{float(op_data[2]):.2f}" if op_data[2] is not None else "0.00",  # Время замера
                        f"{float(op_data[3]):.4f}" if op_data[3] is not None else "0.0000",  # Время на 1 деталь
                        f"{float(op_data[4]):.4f}" if op_data[4] is not None else "0.0000",  # Ставка
                        f"{float(op_data[5]):.2f}" if op_data[5] is not None else "0.00",  # Стоимость
                        str(op_data[6]),  # Сотрудник
                        str(op_data[7]) if op_data[7] is not None else ""  # Утверждённая расценка
                    ]
                    ws_ops.append(row_to_append)
                logger.debug(
                    f"[ИЗДЕЛИЯ_EXCEL] {len(operations) if operations else 0} строк данных добавлено на лист 'Операции'")

                # Форматирование листа "Операции"
                self._format_operations_sheet(ws_ops)
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Операции' отформатирован")

            # --- 7. Создание и заполнение листа "Материалы" ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 7. Создание и заполнение листа 'Материалы'")

            if materials:
                ws_mats = wb.create_sheet(title="Материалы")
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Материалы' создан")

                # Заголовки для материалов
                mats_headers = ["Материал", "Длина (м)", "Ширина (м)", "Количество (шт)", "Стоимость (грн)"]
                ws_mats.append(mats_headers)
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Заголовки добавлены на лист 'Материалы'")

                # Данные материалов
                for mat_data in materials:
                    # Убедимся, что все данные преобразованы в строки или числа
                    row_to_append = [
                        str(mat_data[0]),  # Материал
                        f"{float(mat_data[1]):.3f}" if mat_data[1] is not None else "0.000",  # Длина
                        f"{float(mat_data[2]):.3f}" if mat_data[2] is not None else "0.000",  # Ширина
                        str(mat_data[3]),  # Количество
                        f"{float(mat_data[4]):.2f}" if mat_data[4] is not None else "0.00"  # Стоимость
                    ]
                    ws_mats.append(row_to_append)
                logger.debug(
                    f"[ИЗДЕЛИЯ_EXCEL] {len(materials) if materials else 0} строк данных добавлено на лист 'Материалы'")

                # Форматирование листа "Материалы"
                self._format_materials_sheet(ws_mats)
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Материалы' отформатирован")

            # --- 8. Создание и заполнение листа "Инструкция" ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 8. Создание и заполнение листа 'Инструкция'")

            ws_instr = wb.create_sheet(title="Инструкция")
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Инструкция' создан")

            # Заголовок инструкции
            ws_instr.append(["ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ ФАЙЛА"])
            ws_instr.append([""])  # Пустая строка

            # Текст инструкции
            instruction_text = [
                "Этот файл содержит информацию об изделии и его стоимости.",
                "",
                "Листы:",
                "- 'Информация': Основные данные об изделии.",
                "- 'Операции': Технологические операции и их стоимость.",
                "- 'Материалы': Используемые материалы и их стоимость.",
                "- 'Инструкция': Эта страница.",
                "",
                "ВАЖНО:",
                "- Не изменяйте структуру файла вручную.",
                "- Для изменения данных используйте программу.",
                "- Файл автоматически сохраняется в БД при каждом обновлении.",
                "- При повреждении файла его можно восстановить из БД.",
                "",
                "Дата создания файла: " + pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Версия программы: 1.0"
            ]

            for line in instruction_text:
                ws_instr.append([line])
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Текст инструкции добавлен на лист 'Инструкция'")

            # Форматирование листа "Инструкция"
            self._format_instruction_sheet(ws_instr)
            logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Инструкция' отформатирован")

            # --- 9. Финальная проверка и сохранение ---
            logger.debug("[ИЗДЕЛИЯ_EXCEL] 9. Финальная проверка и сохранение")

            # Убедимся, что лист "Информация" существует и сделаем его активным
            if 'Информация' in wb.sheetnames:
                wb.active = wb['Информация']
                logger.debug("[ИЗДЕЛИЯ_EXCEL] Лист 'Информация' установлен как активный")
            elif wb.sheetnames:
                # Если "Информация" по какой-то причине отсутствует, активируем первый доступный
                wb.active = wb[wb.sheetnames[0]]
                logger.warning(
                    f"[ИЗДЕЛИЯ_EXCEL] Лист 'Информация' не найден. Активирован первый лист: {wb.sheetnames[0]}")
            else:
                # Крайне маловероятный случай
                logger.error("[ИЗДЕЛИЯ_EXCEL] Критическая ошибка: Нет листов в книге!")
                wb.close()
                return False

            # Сохраняем файл
            logger.debug(f"[ИЗДЕЛИЯ_EXCEL] Попытка сохранения книги в '{file_path}'")
            wb.save(file_path)
            wb.close()
            logger.info(f"[ИЗДЕЛИЯ_EXCEL] Файл успешно сохранен в '{file_path}'")

            # Проверим, что файл действительно создан и не пустой
            if os.path.exists(file_path):
                size = os.path.getsize(file_path)
                logger.debug(f"[ИЗДЕЛИЯ_EXCEL] Файл '{file_path}' существует, размер: {size} байт")
                if size == 0:
                    logger.error(f"[ИЗДЕЛИЯ_EXCEL] Файл '{file_path}' создан, но имеет нулевой размер!")
                    return False
            else:
                logger.error(f"[ИЗДЕЛИЯ_EXCEL] Файл '{file_path}' не был создан!")
                return False

            logger.info(f"[ИЗДЕЛИЯ_EXCEL] Сохранение изделия ID {product_id} в Excel завершено успешно.")
            return True
        except Exception as e:
            logger.error(f"[ИЗДЕЛИЯ_EXCEL] Ошибка при сохранении изделия в Excel: {e}", exc_info=True)
            return False

    def _format_info_sheet(self, ws):
        """Форматирование листа 'Информация'"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные
        for row in ws.iter_rows(min_row=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Автоподбор ширины колонок
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 50)  # Ограничиваем максимальную ширину

    def _format_operations_sheet(self, ws):
        """Форматирование листа 'Операции'"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Автоподбор ширины колонок
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 50)  # Ограничиваем максимальную ширину

    def _format_materials_sheet(self, ws):
        """Форматирование листа 'Материалы'"""
        # Заголовки
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Данные
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Автоподбор ширины колонок
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = min(value + 2, 50)  # Ограничиваем максимальную ширину

    def _format_instruction_sheet(self, ws):
        """Форматирование листа 'Инструкция'"""
        # Заголовок
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        # Текст
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Автоподбор ширины колонок
        ws.column_dimensions['A'].width = 80  # Фиксированная ширина для инструкции

    def get_all_products(self):
        """Получение всех изделий"""
        logger.debug("[ИЗДЕЛИЯ] Получение списка всех изделий из БД")
        query = "SELECT id, product_id, article, name, created_date FROM products ORDER BY name"
        return self.db_manager.fetch_all(query)

    def load_product_from_excel(self, file_path):
        """Загрузка изделия из Excel файла"""
        logger.info(f"[ИЗДЕЛИЯ] Загрузка изделия из файла: {file_path}")
        # Реализация загрузки из Excel
        # Пока заглушка
        return None